"""
utils/reports.py
Generates professional PDF and Excel reports, plus cleaned CSV export.
"""
import os
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, Image, PageBreak)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io

from utils.charts import generate_static_heatmap_image, generate_static_missing_chart


def export_cleaned_csv(df: pd.DataFrame, output_path: str) -> str:
    df.to_csv(output_path, index=False)
    return output_path


def export_excel_report(df: pd.DataFrame, profile: dict, quality: dict,
                         insights: dict, output_path: str) -> str:
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    header_fill = PatternFill(start_color="2E5C88", end_color="2E5C88", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws["A1"] = "Cleany - Data Quality & Profiling Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:B1")

    rows = [
        ("Total Rows", profile["total_rows"]),
        ("Total Columns", profile["total_columns"]),
        ("Missing Values (%)", profile["missing_percent"]),
        ("Duplicate Count", profile["duplicate_count"]),
        ("Data Quality Score", quality["score"]),
        ("Quality Grade", quality["grade"]),
    ]
    ws.append([])
    for label, value in rows:
        ws.append([label, value])
    for row in ws.iter_rows(min_row=3, max_row=2 + len(rows), max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    # --- Insights sheet ---
    ws2 = wb.create_sheet("AI Insights")
    ws2["A1"] = "Executive Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = insights.get("executive_summary", "")
    ws2["A2"].alignment = Alignment(wrap_text=True)
    ws2.merge_cells("A2:D2")

    section_row = 4
    for section in ["key_findings", "business_insights", "risks", "recommendations",
                     "data_quality_observations"]:
        ws2.cell(row=section_row, column=1, value=section.replace("_", " ").title()).font = Font(bold=True, size=12)
        section_row += 1
        for item in insights.get(section, []):
            ws2.cell(row=section_row, column=1, value=f"- {item}")
            section_row += 1
        section_row += 1

    # --- Cleaned data sheet ---
    ws3 = wb.create_sheet("Cleaned Data")
    for r in dataframe_to_rows(df.head(2000), index=False, header=True):
        ws3.append(r)
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save(output_path)
    return output_path


def export_pdf_report(df: pd.DataFrame, profile: dict, understanding: dict, quality: dict,
                       cleaning_report: dict, insights: dict, output_path: str,
                       dataset_name: str = "Dataset") -> str:
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                             topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], textColor=colors.HexColor("#2E5C88"))
    heading_style = ParagraphStyle("Heading", parent=styles["Heading2"], textColor=colors.HexColor("#2E5C88"),
                                    spaceBefore=12, spaceAfter=6)
    body_style = styles["BodyText"]

    elements = []
    elements.append(Paragraph("Cleany - Automated Data Report", title_style))
    elements.append(Paragraph(f"Dataset: {dataset_name}", body_style))
    elements.append(Spacer(1, 12))

    # Executive Summary
    elements.append(Paragraph("Executive Summary", heading_style))
    elements.append(Paragraph(insights.get("executive_summary", ""), body_style))

    # Profile table
    elements.append(Paragraph("Dataset Profile", heading_style))
    profile_data = [
        ["Metric", "Value"],
        ["Total Rows", profile["total_rows"]],
        ["Total Columns", profile["total_columns"]],
        ["Missing Values (%)", f"{profile['missing_percent']}%"],
        ["Duplicate Count", profile["duplicate_count"]],
        ["Duplicates Removed", cleaning_report.get("duplicates_removed", 0)],
        ["Domain", understanding.get("domain", "N/A")],
    ]
    t = Table(profile_data, colWidths=[8 * cm, 8 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5C88")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(t)

    # Quality score
    elements.append(Paragraph("Data Quality Score", heading_style))
    elements.append(Paragraph(f"<b>{quality['score']} / 100 ({quality['grade']})</b>", body_style))
    qb = quality["breakdown"]
    qb_data = [["Penalty Type", "Value"]] + [[k.replace("_", " ").title(), v] for k, v in qb.items()]
    t2 = Table(qb_data, colWidths=[8 * cm, 8 * cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4C72B0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(t2)

    # Charts
    heatmap_img = generate_static_heatmap_image(df)
    if heatmap_img:
        elements.append(Paragraph("Correlation Heatmap", heading_style))
        elements.append(Image(io.BytesIO(heatmap_img), width=14 * cm, height=11 * cm))

    missing_img = generate_static_missing_chart(df)
    if missing_img:
        elements.append(Paragraph("Missing Values Chart", heading_style))
        elements.append(Image(io.BytesIO(missing_img), width=14 * cm, height=9 * cm))

    elements.append(PageBreak())

    # AI Insights sections
    for section_key, section_title in [
        ("key_findings", "Key Findings"),
        ("business_insights", "Business Insights"),
        ("risks", "Risks"),
        ("recommendations", "Recommendations"),
        ("data_quality_observations", "Data Quality Observations"),
    ]:
        elements.append(Paragraph(section_title, heading_style))
        for item in insights.get(section_key, []):
            elements.append(Paragraph(f"&bull; {item}", body_style))
        elements.append(Spacer(1, 6))

    doc.build(elements)
    return output_path
